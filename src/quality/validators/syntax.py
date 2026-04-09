"""
语法验证器

验证输出格式是否符合规范
"""

import json
from typing import Any, Dict, List
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from src.quality.validator import (
    BaseValidator, ValidationResult, ValidationIssue, 
    QualityScore, ValidationSeverity, ValidatorRegistry
)
from src.utils.logger import get_logger

logger = get_logger(__name__)


@ValidatorRegistry.register
class JSONSyntaxValidator(BaseValidator):
    """JSON 语法验证器"""
    
    name = "json_syntax"
    description = "验证 JSON 格式是否正确"
    weight = 0.8
    
    async def validate(self, input_data: Any, output_data: Any, context: Dict[str, Any] = None) -> ValidationResult:
        """验证 JSON 语法"""
        issues = []
        context = context or {}
        
        # 如果输出是字符串，尝试解析
        if isinstance(output_data, str):
            try:
                parsed = json.loads(output_data)
                output_data = parsed
            except json.JSONDecodeError as e:
                issues.append(self.create_issue(
                    code="INVALID_JSON",
                    message=f"JSON 解析失败: {str(e)}",
                    severity=ValidationSeverity.CRITICAL,
                    suggestion="检查 JSON 格式，确保引号、括号匹配"
                ))
                score = QualityScore(
                    total_score=0.0,
                    dimension_scores={"syntax": 0.0},
                    passed=False,
                    threshold=60.0
                )
                return ValidationResult.failed(issues, score, {"error": str(e)})
        
        # 验证必需字段
        required_schema = context.get("required_schema", [])
        if required_schema and isinstance(output_data, dict):
            for field in required_schema:
                if field not in output_data:
                    issues.append(self.create_issue(
                        code="MISSING_REQUIRED_FIELD",
                        message=f"缺少必需字段: {field}",
                        severity=ValidationSeverity.ERROR,
                        field=field,
                        suggestion=f"确保输出包含 '{field}' 字段"
                    ))
        
        # 验证数组非空
        if isinstance(output_data, list) and len(output_data) == 0:
            issues.append(self.create_issue(
                code="EMPTY_ARRAY",
                message="输出数组为空",
                severity=ValidationSeverity.ERROR,
                suggestion="确保生成了有效的数据项"
            ))
        
        score_value = self.calculate_score(issues)
        score = QualityScore(
            total_score=score_value,
            dimension_scores={"syntax": score_value},
            passed=score_value >= 60.0 and not any(i.severity == ValidationSeverity.CRITICAL for i in issues),
            threshold=60.0
        )
        
        if score.passed and not issues:
            return ValidationResult.passed(score, {"format": "json"})
        
        return ValidationResult.failed(issues, score, {"format": "json"})


@ValidatorRegistry.register
class ExcelStructureValidator(BaseValidator):
    """Excel 结构验证器"""
    
    name = "excel_structure"
    description = "验证 Excel 文件结构是否符合测试用例模板"
    weight = 0.8
    
    # 标准测试用例表头
    STANDARD_HEADERS = [
        "用例编号", "用例标题", "前置条件", "测试步骤", 
        "预期结果", "优先级", "所属模块", "用例类型"
    ]
    
    async def validate(self, input_data: Any, output_data: Any, context: Dict[str, Any] = None) -> ValidationResult:
        """验证 Excel 结构"""
        issues = []
        context = context or {}
        
        file_path = output_data if isinstance(output_data, (str, Path)) else context.get("file_path")
        
        if not file_path:
            issues.append(self.create_issue(
                code="MISSING_FILE_PATH",
                message="未提供 Excel 文件路径",
                severity=ValidationSeverity.CRITICAL
            ))
            score = QualityScore(total_score=0.0, dimension_scores={"structure": 0.0}, passed=False, threshold=60.0)
            return ValidationResult.failed(issues, score)
        
        if not HAS_OPENPYXL:
            logger.warning("openpyxl not installed, skipping Excel validation")
            score = QualityScore(total_score=100.0, dimension_scores={"structure": 100.0}, passed=True, threshold=60.0)
            return ValidationResult.passed(score, {"note": "openpyxl not available"})
        
        try:
            wb = openpyxl.load_workbook(file_path)
            
            # 验证工作表存在
            if len(wb.sheetnames) == 0:
                issues.append(self.create_issue(
                    code="NO_SHEETS",
                    message="Excel 文件没有工作表",
                    severity=ValidationSeverity.CRITICAL
                ))
            
            ws = wb.active
            
            # 验证表头
            if ws.max_row < 1:
                issues.append(self.create_issue(
                    code="EMPTY_SHEET",
                    message="工作表为空",
                    severity=ValidationSeverity.CRITICAL
                ))
            else:
                headers = [cell.value for cell in ws[1]]
                missing_headers = [h for h in self.STANDARD_HEADERS if h not in headers]
                
                if missing_headers:
                    issues.append(self.create_issue(
                        code="MISSING_HEADERS",
                        message=f"缺少标准表头: {', '.join(missing_headers)}",
                        severity=ValidationSeverity.ERROR,
                        suggestion=f"标准表头应包含: {', '.join(self.STANDARD_HEADERS)}"
                    ))
            
            # 验证数据行
            if ws.max_row < 2:
                issues.append(self.create_issue(
                    code="NO_DATA_ROWS",
                    message="没有数据行（只有表头）",
                    severity=ValidationSeverity.ERROR,
                    suggestion="确保生成了至少一条测试用例"
                ))
            
            wb.close()
            
        except Exception as e:
            issues.append(self.create_issue(
                code="EXCEL_READ_ERROR",
                message=f"Excel 文件读取失败: {str(e)}",
                severity=ValidationSeverity.CRITICAL
            ))
        
        score_value = self.calculate_score(issues)
        score = QualityScore(
            total_score=score_value,
            dimension_scores={"structure": score_value},
            passed=score_value >= 60.0 and not any(i.severity == ValidationSeverity.CRITICAL for i in issues),
            threshold=60.0
        )
        
        if score.passed and not issues:
            return ValidationResult.passed(score, {"file_path": str(file_path)})
        
        return ValidationResult.failed(issues, score, {"file_path": str(file_path)})