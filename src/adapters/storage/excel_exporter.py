"""
Excel导出器

将测试用例导出为Excel格式
"""

from pathlib import Path
from typing import List, Dict, Any, TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from src.core.config import Config
from src.utils.logger import get_logger

if TYPE_CHECKING:
    from src.skills.test_case_generator.models import TestSuite, TestCase

logger = get_logger(__name__)


class ExcelExporter:
    """
    Excel导出器
    
    使用示例：
        ```python
        exporter = ExcelExporter(config)
        await exporter.export_test_suite(test_suite, "output.xlsx")
        ```
    """
    
    def __init__(self, config: Config):
        """
        初始化导出器
        
        Args:
            config: 配置对象
        """
        self.config = config
    
    async def export_test_suite(
        self, 
        test_suite: "TestSuite",
        output_path: str
    ) -> str:
        """
        导出测试套件到Excel - 标准测试用例格式
        
        Args:
            test_suite: 测试套件
            output_path: 输出文件路径
        
        Returns:
            导出文件路径
        """
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"
        
        # 标准测试用例表格格式（与web/api.py保持一致）
        headers = ['用例编号', '所属模块', '用例标题', '前置条件', '测试步骤', '预期结果', '优先级', '执行结果', '备注']
        
        # 定义样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="C4A77D", end_color="C4A77D", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        data_font = Font(size=10)
        data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        priority_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='d4c8b8'),
            right=Side(style='thin', color='d4c8b8'),
            top=Side(style='thin', color='d4c8b8'),
            bottom=Side(style='thin', color='d4c8b8')
        )
        
        # 写入标题行
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 设置表头行高
        ws.row_dimensions[1].height = 30
        
        # 优先级颜色映射
        priority_colors = {
            'P0': 'FF6B6B',  # 红色
            'P1': 'FFB347',  # 橙色
            'P2': '87CEEB',  # 蓝色
            'P3': '90EE90',  # 绿色
        }
        
        # 写入数据
        for row_idx, test_case in enumerate(test_suite.test_cases, 2):
            # 格式化测试步骤（带序号换行）
            steps_text = ""
            expected_text = ""
            if test_case.steps:
                steps_lines = []
                expected_lines = []
                for i, step in enumerate(test_case.steps, 1):
                    steps_lines.append(f"{i}. {step.action}")
                    expected_lines.append(f"{i}. {step.expected_result}")
                steps_text = "\n".join(steps_lines)
                expected_text = "\n".join(expected_lines)
            
            # 前置条件
            preconditions_text = "\n".join(test_case.preconditions) if test_case.preconditions else ""
            
            row_data = [
                test_case.id,                    # 用例编号
                test_case.module or "默认模块",   # 所属模块
                test_case.title,                  # 用例标题
                preconditions_text,               # 前置条件
                steps_text,                       # 测试步骤
                expected_text,                    # 预期结果
                test_case.priority,               # 优先级
                "",                               # 执行结果（空白待填写）
                test_case.notes or "",            # 备注
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                
                # 优先级列居中并着色
                if col_idx == 7:  # 优先级列
                    cell.alignment = priority_alignment
                    priority = test_case.priority.upper()
                    if priority in priority_colors:
                        cell.fill = PatternFill(start_color=priority_colors[priority], 
                                               end_color=priority_colors[priority], 
                                               fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF', size=10)
                # 用例编号居中
                elif col_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = data_alignment
            
            # 设置行高（根据步骤数量自动调整）
            step_count = len(test_case.steps) if test_case.steps else 1
            ws.row_dimensions[row_idx].height = max(40, step_count * 15 + 10)
        
        # 设置列宽
        ws.column_dimensions['A'].width = 12  # 用例编号
        ws.column_dimensions['B'].width = 15  # 所属模块
        ws.column_dimensions['C'].width = 30  # 用例标题
        ws.column_dimensions['D'].width = 25  # 前置条件
        ws.column_dimensions['E'].width = 45  # 测试步骤
        ws.column_dimensions['F'].width = 35  # 预期结果
        ws.column_dimensions['G'].width = 10  # 优先级
        ws.column_dimensions['H'].width = 10  # 执行结果
        ws.column_dimensions['I'].width = 20  # 备注
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 保存
        wb.save(path)
        logger.info(f"Test suite exported to: {path}")
        
        return str(path)
    
    async def export_test_cases_to_csv(
        self,
        test_cases: List["TestCase"],
        output_path: str
    ) -> str:
        """
        导出测试用例到CSV
        
        Args:
            test_cases: 测试用例列表
            output_path: 输出文件路径
        
        Returns:
            导出文件路径
        """
        import pandas as pd
        
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        
        # 转换为DataFrame - 使用标准格式
        data = []
        for tc in test_cases:
            steps_text = ""
            expected_text = ""
            if tc.steps:
                steps_lines = [f"{s.step_number}. {s.action}" for s in tc.steps]
                expected_lines = [f"{s.step_number}. {s.expected_result}" for s in tc.steps]
                steps_text = "\n".join(steps_lines)
                expected_text = "\n".join(expected_lines)
            
            data.append({
                '用例编号': tc.id,
                '所属模块': tc.module or '默认模块',
                '用例标题': tc.title,
                '前置条件': "\n".join(tc.preconditions) if tc.preconditions else "",
                '测试步骤': steps_text,
                '预期结果': expected_text,
                '优先级': tc.priority,
                '执行结果': '',
                '备注': tc.notes or '',
            })
        
        df = pd.DataFrame(data)
        
        # 保存
        df.to_csv(path, index=False, encoding='utf-8-sig')
        logger.info(f"Test cases exported to CSV: {path}")
        
        return str(path)